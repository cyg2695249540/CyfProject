# !/usr/bin/env Python3
# -*- coding: utf-8 -*-
# @FILE     : excel_data2.py
# @Author   : Pluto.
# @Time     : 2020/9/10 14:08
from openpyxl import Workbook, load_workbook


class PracticeExcel:
    def create(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"
        self.height = [180, 160, 170, 155]
        weight = [66, 50, 40, 30]
        for i in range(len(self.height)):
            ws.cell(row=2 + i, column=1, value=self.height[i])
            ws.cell(row=2 + i, column=2, value=weight[i])
        wb.save("sample.xlsx")

    def health_weight(self):
        ld = load_workbook(filename = "sample.xlsx")
        sheet = ld.active
        sheet["C1"] = "备注"
        for i in range(len(self.height)):
            height = sheet.cell(row=2 + i, column=1).value
            weight = sheet.cell(row=2 + i, column=2).value
            health_w = (height-70)*0.6
            if weight == health_w:
                print("这是健康的体重",weight)
                sheet.cell(row=2 + i, column=3, value="健康体重")

        ld.save("sample1.xlsx")

pe = PracticeExcel()
pe.create()
pe.health_weight()