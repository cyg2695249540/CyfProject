# !/usr/bin/env Python3
# -*- coding: utf-8 -*-
# @FILE     : excel_data.py
# @Author   : Pluto.
# @Time     : 2020/9/10 13:31
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "身高"
ws['B1'] = "体重"
height = [180,160,170,155]
weight = [60,50,40,30]
for i in range(len(height)):
    ws.cell(row = 2 + i, column = 1, value = height[i])
    ws.cell(row = 2 + i, column = 2, value = weight[i])
wb.save("sample.xlsx")